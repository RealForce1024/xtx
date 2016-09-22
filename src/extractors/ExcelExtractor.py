#!/usr/bin/env python
# -*- coding: utf-8  -*-

import os.path
import pandas as pd

class ExcelExtractor(object):

	def __init__(self):
		pass

	def __readExcel2003(self, filepath, targetSheetIndex = None, \
		targetSheetName = None, headerRows = 1):
		import xlrd

		tabledata = []
		wb = xlrd.open_workbook(filepath)

		# get worksheet
		ws = None
		if not targetSheetName is None:
			ws = wb.sheet_by_name(targetSheetName)
		else:
			if not targetSheetIndex is None:
				ws = wb.sheet_by_index(targetSheetIndex)
			else:
				raise ValueError("The targetSheet arguments must be specified one:  targetSheetIndex=%s, targetSheetName=%s" % (targetSheetIndex, targetSheetName))

		# get cell data
		for rowIndex in range(0, ws.nrows):
			row = ws.row(rowIndex)
			rowdata = []
			for colIndex in range(0, ws.ncols):
				cell = row[colIndex]
				cellval = cell.value
				rowdata.append(cellval)
			tabledata.append(rowdata)
			
		return pd.DataFrame(data = tabledata[headerRows:], columns = tabledata[:headerRows][0])

	def __readExcel2007(self, filepath, targetSheetIndex = None, \
		targetSheetName = None, headerRows = 1):
		from openpyxl.workbook import Workbook
		from openpyxl import load_workbook
		
		tabledata = []
		wb = load_workbook(filename = filepath)
		
		# get worksheet
		ws = None
		if not targetSheetName is None:
			ws = wb.get_sheet_by_name(targetSheetName)
		else:
			if not targetSheetIndex is None:
				ws = wb.get_sheet_by_name(wb.get_sheet_names()[targetSheetIndex])
			else:
				raise ValueError("The targetSheet arguments must be specified one:  targetSheetIndex=%s, targetSheetName=%s" % (targetSheetIndex, targetSheetName))
		
		# get cell data
		for rowNum in range(1, ws.max_row + 1):
			rowdata = []
			for colNum in range(1, ws.max_column + 1):
				cellval = ws.cell(row = rowNum, column= colNum).value
				rowdata.append(cellval)
			tabledata.append(rowdata)
		
		return pd.DataFrame(data = tabledata[headerRows:], columns = tabledata[:headerRows][0])

	def extract(self, filepath, targetSheetIndex = None, targetSheetName = None, headerRows = 1):
		"""
		Extract data from the excel file.

		Args:
			filepath: the path of target file.
			targetSheetIndex: default 0. the index of target sheet.
			targetSheetName: default None. the name of target sheet.
			headerRows: default 1. the rows count of header.

		Returns:
			return the pandas.DataFrame of rows.

		Raises:
			None
		"""

		if(not os.path.exists(filepath)):
			raise FileNotFoundError(filepath)

		df = None
		extName = filepath.split(".")[-1]
		if extName == "xls":
			df = self.__readExcel2003(filepath = filepath, targetSheetIndex = targetSheetIndex, \
			targetSheetName = targetSheetName, headerRows = headerRows)
		elif extName == "xlsx":
			df = self.__readExcel2007(filepath = filepath, targetSheetIndex = targetSheetIndex, \
			targetSheetName = targetSheetName, headerRows = headerRows)
		else:
			raise ValueError("Unsupported file type: %s" % extName)
		return df

if __name__ == "__main__":
	filepath = os.path.abspath(r"../../test/read_test/data/test.xls")
	er = ExcelExtractor()
	print(er.extract.__doc__)
	df = er.extract(filepath = filepath, targetSheetIndex = 0)
	print(df)
	filepath = os.path.abspath(r"../../test/read_test/data/test.xlsx")
	df = er.extract(filepath = filepath, targetSheetIndex = 0)
	print(df)