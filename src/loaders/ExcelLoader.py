#!/usr/bin/env python
# -*- coding: utf-8  -*-

import os.path

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

class ExcelLoader(object):
	
	def __init__(self, srcData, targetPath, targetSheet, recreateFile = False):
		self.srcData = srcData
		self.targetPath = targetPath
		self.targetSheet = targetSheet
		self.recreateFile = recreateFile
		
	def __recreateFile(self):
		if os.path.exists(self.targetPath):
			os.remove(self.targetPath)
		wb = Workbook()
		wb.create_sheet(0)
		wb.save(self.targetPath)
		
	def write(self):
		if not os.path.exists(self.targetPath):
			self.__recreateFile()
		else:
			if self.recreateFile == True:
				self.__recreateFile()
		
		wb = load_workbook(self.targetPath)
		ws = None
		if isinstance(self.targetSheet, str):
			ws = wb.get_sheet_by_name(self.targetSheet) if self.targetSheet in wb.sheetnames else wb.create_sheet(title = self.targetSheet)
		elif isinstance(self.targetSheet, int):
			sheetCount = len(wb.sheetnames)
			ws = wb.create_sheet(index = self.targetSheet) if self.targetSheet < sheetCount else wb.create_sheet(index = sheetCount)
		else:
			raise ValueError("Unsupported arg type: %s" % self.targetSheet)
		
		for rowIndex in range(0, len(self.srcData)):
			rowdata = self.srcData[rowIndex]
			for colIndex in range(0, len(rowdata)):
				celldata = rowdata[colIndex]
				ws.cell(row = rowIndex + 1, column = colIndex + 1).value = celldata
		wb.save(self.targetPath)
		
if __name__ == "__main__":
	testdata = [["A","B","C"],["D","E","F"]]
	ew = ExcelLoader(testdata, r"D:\merge.xlsx", targetSheet = 0, recreateFile = True)
	ew.write()
	print("OK")