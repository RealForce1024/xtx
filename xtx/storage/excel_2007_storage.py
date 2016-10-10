#!/usr/bin/env python
# -*- coding: utf-8  -*-

import os.path

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

from xtx.storage.bin_file_storage import BinFileStorage
from xtx.storage.exceptions import (StorageExistsError
	, StorageNotFoundError
	, UnmatchExtensionError
	, ArgumentsAbsenceError)


class Excel2007Storage(BinFileStorage):
	"""
	Excel 2007 format Storage
	"""
	

	def __init__(self, filepath = None, sheetIndex = -1, sheetName = None):
		super().__init__(filepath)
		self.sheetIndex = sheetIndex
		self.sheetName = sheetName


	def __get_worksheet(self, workbook_toread):
		ws = None
		if not self.sheetName is None:
			ws = workbook_toread.get_sheet_by_name(self.sheetName)
		else:
			if not self.sheetIndex is None:
				ws = workbook_toread.get_sheet_by_name(workbook_toread.get_sheet_names()[self.sheetIndex])
			else:
				raise ArgumentsAbsenceError("The targetSheet arguments must be specified one:  sheetIndex=%s, sheetName=%s" % \
					(self.sheetIndex, self.sheetName))
		return ws


	def create(self, force = False):
		if self.exists():
			if force == False:
				raise StorageExistsError(self.filepath)
			else:
				self.remove(force = True)

		wb = Workbook()
		createSheetName = self.sheetName if self.sheetName is not None else ("SHEET_" + str(self.sheetIndex))
		wb.create_sheet(title =createSheetName)
		wb.save(self.filepath)


	def clear(self, force = False):
		wb = load_workbook(filename = self.filepath)
		ws = self.__get_worksheet(wb)
		wb.remove(ws)
		createSheetName = self.sheetName if self.sheetName is not None else ("SHEET_" + str(self.sheetIndex))
		wb.create_sheet(title =createSheetName)
		wb.save(self.filepath)
		

	def read(self, limit =  -1):
		tabledata = []
		wb = load_workbook(filename = self.filepath)
		ws = self.__get_worksheet(wb)
		
		# get cell data
		for rowNum in range(1, ws.max_row + 1):
			if limit > -1 and rowNum > limit:
				break

			rowdata = []
			for colNum in range(1, ws.max_column + 1):
				cellval = ws.cell(row = rowNum, column= colNum).value
				rowdata.append(cellval)
			tabledata.append(rowdata)
		
		return tabledata

	def write(self, data, overwrite = False):
		wb = load_workbook(filename = self.filepath)
		ws = self.__get_worksheet(wb)
		for rowIndex in range(0, len(data)):
			rowdata = data[rowIndex]
			for colIndex in range(0, len(rowdata)):
				celldata = rowdata[colIndex]
				ws.cell(row = rowIndex + 1, column = colIndex + 1).value = celldata
		wb.save(self.filepath)

