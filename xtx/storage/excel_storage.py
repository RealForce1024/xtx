#!/usr/bin/env python
# -*- coding: utf-8  -*-

import os.path

from xtx.storage.bin_file_storage import BinFileStorage
from xtx.storage.exceptions import (StorageExistsError
	, StorageNotFoundError
	, UnmatchExtensionError
	, ArgumentsAbsenceError)

class Excel2003Storage(BinFileStorage):
	"""
	Excel 2003 format Storage
	"""
	
	import xlrd
	import xlwt

	def __init__(self, filepath = None, sheetIndex = -1, sheetName = None):
		super().__init__(filepath)
		self.sheetIndex = sheetIndex
		self.sheetName = sheetName


	def __get_worksheet(self, workbook_toread):
		ws = None
		if self.sheetName is not None:
			ws = workbook_toread.sheet_by_name(self.sheetName)
		else:
			if self.sheetIndex is not -1:
				ws = workbook_toread.sheet_by_index(self.sheetIndex)
			else:
				raise ArgumentsAbsenceError("The targetSheet arguments must be specified one:  sheetIndex=%s, sheetName=%s" % \
					(self.sheetIndex, self.sheetName))
		return ws


	def create(self, force = False):
		if self.exists():
			if force == False:
				raise StorageExistsError(self.filepath)
			else:
				self.remove(force = True)

		wb_w = xlwt.Workbook(encoding='utf-8', style_compression=0)
		createSheetName = self.sheetName if self.sheetName is not None else ("SHEET_" + self.sheetIndex)
		sheet = wb_w.add_sheet(createSheetName, cell_overwrite_ok = True)
		wb_w.save(self.filepath)


	def clear(self, force = False):
		if not self.exists():
			if force == False:
				raise StorageNotFoundError(self.filepath)
		
		raise NotImplementedError()


	def read(self, limit =  -1):
		tabledata = []
		wb = xlrd.open_workbook(self.filepath)
		ws = self.__get_worksheet(wb)

		# get cell data
		for rowIndex in range(0, ws.nrows):
			if (rowIndex + 1) > limit:
				break

			row = ws.row(rowIndex)
			rowdata = []
			for colIndex in range(0, ws.ncols):
				cell = row[colIndex]
				cellval = cell.value
				rowdata.append(cellval)
			tabledata.append(rowdata)

		return tabledata

	def write(self, data, overwrite = False):
		raise NotImplementedError()






class Excel2007Storage(BinFileStorage):
	"""
	Excel 2007 format Storage
	"""


	from openpyxl.workbook import Workbook
	from openpyxl import load_workbook

	def __init__(self, filepath = None, sheetIndex = -1, sheetName = None):
		super().__init__(filepath)
		self.sheetIndex = sheetIndex
		self.sheetName = sheetName


	def __get_worksheet(self, workbook_toread):
		ws = None
		if not self.sheetName is None:
			ws = workbook_toread.get_sheet_by_name(self.sheetName)
		else:
			if not self.sheetIndex is None:
				ws = workbook_toread.get_sheet_by_name(wb.get_sheet_names()[self.sheetIndex])
			else:
				raise ArgumentsAbsenceError("The targetSheet arguments must be specified one:  sheetIndex=%s, sheetName=%s" % \
					(self.sheetIndex, self.sheetName))
		return ws


	def create(self, force = False):
		if self.exists():
			if force == False:
				raise StorageExistsError(self.filepath)
			else:
				self.remove(force = True)

		wb = Workbook()
		createSheetName = self.sheetName if self.sheetName is not None else ("SHEET_" + self.sheetIndex)
		wb.create_sheet(title =createSheetName)
		wb.save(self.filepath)


	def clear(self, force = False):
		wb = load_workbook(filename = self.filepath)
		ws = self.__get_worksheet(wb)
		wb.remove(ws)
		createSheetName = self.sheetName if self.sheetName is not None else ("SHEET_" + self.sheetIndex)
		wb.create_sheet(title =createSheetName)
		wb.save(self.filepath)
		

	def read(self, limit =  -1):
		tabledata = []
		wb = load_workbook(filename = self.filepath)
		ws = self.__get_worksheet(wb)
		
		# get cell data
		for rowNum in range(1, ws.max_row + 1):
			if rowNum > limit:
				break

			rowdata = []
			for colNum in range(1, ws.max_column + 1):
				cellval = ws.cell(row = rowNum, column= colNum).value
				rowdata.append(cellval)
			tabledata.append(rowdata)
		
		return tabledata

	def write(self, data, overwrite = False):
		raise NotImplementedError()



class ExcelStorage(BinFileStorage):

	def __init__(self, filepath = None, sheetIndex = -1, sheetName = None):
		super().__init__(filepath)
		self.sheetIndex = sheetIndex
		self.sheetName = sheetName

		ext = self.filepath.split(".")[-1]
		if ext == "xls":
			pass
		elif ext == "xlsx":
			pass
		else:
			raise UnmatchExtensionError(self.filepath)

	
	def getExtension(self):
		return self.filepath.split(".")[-1]

	
	def __create_excel2003(self):
		pass


	def __create_excel2007(self):
		pass

		
	def __read_excel2003(self, limit = -1):
		pass

		

	def __read_excel2007(self, limit = -1):
		pass
		
		

	def __write_excel2003(self, data, overwrite = False):
		raise NotImplementedError(self.filepath)

	def __write_excel2007(self, data, overwrite = False):
		raise NotImplementedError(self.filepath)



	def create(self, force = False):
		raise NotImplementedError(self.filepath)


	def clear(self, force = False):
		raise NotImplementedError(self.filepath)

		
	def write(self, data, overwrite = False):
		if not self.exists(self.filepath):
			raise StorageNotFoundError(self.filepath)

		ext = self.getExtension()
		if ext == "xls":
			self.__write_excel2003(data = data, overwrite = overwrite)
		elif ext == "xlsx":
			self.__write_excel2007(data = data, overwrite = overwrite)
		else:
			raise UnmatchExtensionError(self.filepath)


	def read(self, limit = -1):
		if not self.exists(self.filepath):
			raise StorageNotFoundError(self.filepath)
		
		data = None
		ext = self.getExtension()
		if ext == "xls":
			data = self.__read_excel2003(limit = limit)
		elif ext == "xlsx":
			data = self.__read_excel2007(limit = limit)
		else:
			raise UnmatchExtensionError(self.filepath)
		return data
	
